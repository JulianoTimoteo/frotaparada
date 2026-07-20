"""
Script de importacao da planilha "Programacao Geral 1.xlsx"
Le a planilha, filtra por Gestor, gera JSON para upload via frontend
"""

import openpyxl
from datetime import datetime
import json
import os
import glob

# === CONFIG ===
DIR = os.path.dirname(__file__)
# Encontra a planilha independente de acentos
planilhas = glob.glob(os.path.join(DIR, "Programa*Geral*1*.xlsx")) or glob.glob(os.path.join(DIR, "Programa*1*.xlsx"))
CAMINHO_PLANILHA = planilhas[0] if planilhas else os.path.join(DIR, "Programacao Geral 1.xlsx")
CAMINHO_JSON = os.path.join(os.path.dirname(__file__), "preventivas_dados.json")
GESTORES_FILTRO = ["Reginaldo Mantovani"]

# Merge de subclasses equivalentes
SUBCLASSE_MERGE = {
    "SEMI-REBOQUES CANAVIEIROS": "REBOQUES CANAVIEIROS",
    "REBOQUE CANAVIEIRO 4 EIXOS": "REBOQUES CANAVIEIROS",
    "REBOQUE CANAVIEIRO 2 EIXOS": "REBOQUES CANAVIEIROS",
    "TRANSPORTE DE ÁGUA (1)": "BOMBEIRO CTT",
}

# Mapeamento de classes para grupos
MAPA_CLASSE_PARA_GRUPO = {}
CLASSE_OUTROS = "demais"

# Mapeamento normalizado (case-insensitive, acentos)
MAPAS_NORMALIZADOS = {
    "CAMINHOES": "caminhoes",
    "IMPLEMENTO RODOVIARIO": "implemento-rod",
    "TRATOR DE PNEUS": "tratores",
    "VEICULOS LEVES": "leves",
    "COLHEDORA": "colhedoras",
    "PA CARREGADEIRA": "pa",
}


def classe_para_grupo(classe):
    if not classe:
        return CLASSE_OUTROS
    # Remove acentos e normaliza
    nome = classe.strip().upper()
    nome = nome.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    nome = nome.replace('Ã', 'A').replace('Õ', 'O').replace('Â', 'A')
    return MAPAS_NORMALIZADOS.get(nome, CLASSE_OUTROS)


def data_para_str(data):
    if data is None:
        return ""
    if isinstance(data, datetime):
        return data.strftime("%Y-%m-%d")
    return str(data).strip()


def construir_documento(row):
    frota_raw = row[0]
    gestor = str(row[1]).strip() if row[1] else ""
    usuario = str(row[2]).strip() if row[2] else ""
    classe = str(row[3]).strip() if row[3] else ""
    subclasse = str(row[4]).strip() if row[4] else ""
    modelo = str(row[5]).strip() if row[5] else ""
    dia = str(row[6]).strip() if row[6] else ""
    data_prog = data_para_str(row[7])
    manutencao = str(row[8]).strip() if row[8] else ""

    if isinstance(frota_raw, float):
        frota_str = str(int(frota_raw))
    else:
        frota_str = str(frota_raw).strip()

    # Aplica merge de subclasses
    subclasse_merged = SUBCLASSE_MERGE.get(subclasse, subclasse)

    grupo = classe_para_grupo(classe)

    return {
        "frota": frota_str,
        "gestor": gestor,
        "usuario": usuario,
        "classe": classe,
        "subclasse": subclasse,
        "subclasse_merged": subclasse_merged,
        "modelo": modelo,
        "dia": dia,
        "data_programacao": data_prog,
        "manutencao_programada": manutencao,
        "grupo": grupo,
        "status": "aberto",
        "hora_inicio": "",
        "data_encerramento": "",
        "ultima_atualizacao": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    }


def main():
    print("=" * 60)
    print("IMPORTADOR DE PREVENTIVAS -> JSON")
    print("Planilha:", CAMINHO_PLANILHA)
    print("Gestores:", ", ".join(GESTORES_FILTRO))
    print("=" * 60)

    if not os.path.exists(CAMINHO_PLANILHA):
        print("ERRO: Planilha nao encontrada:", CAMINHO_PLANILHA)
        # Tenta encontrar o arquivo com acentos
        dir_path = os.path.dirname(__file__)
        for f in os.listdir(dir_path):
            if 'Programacao' in f or 'Programa' in f:
                print("Alternativa encontrada:", f)
        return

    wb = openpyxl.load_workbook(CAMINHO_PLANILHA, data_only=True)

    # Encontra a planilha correta
    sheet_name = None
    for name in wb.sheetnames:
        if 'PROGRAMACAO' in name.upper() or 'PROGRAMA' in name.upper():
            sheet_name = name
            break
    if not sheet_name:
        print("Planilhas disponiveis:", wb.sheetnames)
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]
    print("Aba usada:", sheet_name)
    print("Total de linhas:", ws.max_row - 1)

    documentos = []
    total_filtrados = 0
    grupos = {}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if row[0] is None:
            continue

        gestor = str(row[1]).strip() if row[1] else ""
        usuario = str(row[2]).strip() if row[2] else ""

        # Filtro: gestor + usuario = Reginaldo Mantovani (CTT)
        if gestor not in GESTORES_FILTRO or usuario not in GESTORES_FILTRO:
            continue

        total_filtrados += 1
        doc = construir_documento(row)
        documentos.append(doc)
        grupo = doc["grupo"]
        grupos[grupo] = grupos.get(grupo, 0) + 1

    # Salva JSON
    with open(CAMINHO_JSON, "w", encoding="utf-8") as f:
        json.dump(documentos, f, ensure_ascii=False, indent=2)

    print("\nRESUMO:")
    print("  Registros filtrados:", total_filtrados)
    print("  JSON gerado:", CAMINHO_JSON)
    print("  Tamanho:", os.path.getsize(CAMINHO_JSON), "bytes")
    print("\nDistribuicao por grupo:")
    for grupo, qtd in sorted(grupos.items(), key=lambda x: -x[1]):
        print(f"  {grupo}: {qtd}")


if __name__ == "__main__":
    main()
